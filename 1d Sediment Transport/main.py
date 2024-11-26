import os
import pandas as pd
from grains import GrainReader
from hec import *
from mpm import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import matplotlib.pyplot as plt

def get_char_grain_size(file_name=str, D_char=str):
    grain_info = GrainReader(file_name)
    return grain_info.size_classes["size"][D_char]


def add_header_background_color(file_path, sheet_name="Sheet1"):
    try:
        # Open the Excel workbook
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]

        # Define a fill for the background color (e.g., light blue)
        header_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

        # Apply the fill to the header row (assumes headers are in the first row)
        for cell in sheet[1]:
            cell.fill = header_fill

        # Save the workbook
        wb.save(file_path)
        wb.close()
    except Exception as e:
        print(f"Failed to apply background color: {e}")


def calculate_mpm(hec_df, D_char):
    # create dictionary with relevant information about bed load transport with void lists
    mpm_dict = {
            "River Sta": [],
            "Scenario": [],
            "Q (m3/s)": [],
            "Phi (-)": [],
            "Qb (kg/s)": []
    }

    # extract relevant hydraulic data from HEC-RAS output file
    Froude = hec_df["Froude # Chl"]
    h = hec_df["Hydr Depth"]
    Q = hec_df["Q Total"]
    Rh = hec_df["Hydr Radius"]
    Se = hec_df["E.G. Slope"]
    u = hec_df["Vel Chnl"]

    for i, sta in enumerate(list(hec_df["River Sta"])):
        if not str(sta).lower() == "nan":
            logging.info("PROCESSING PROFILE {0} FOR SCENARIO {1}".format(str(hec_df["River Sta"][i]), str(hec_df["Profile"][i])))
            mpm_dict["River Sta"].append(hec_df["River Sta"][i])
            mpm_dict["Scenario"].append(hec_df["Profile"][i])
            section_mpm = MPM(grain_size=D_char,
                              Froude=Froude[i],
                              water_depth=h[i],
                              velocity=u[i],
                              Q=Q[i],
                              hydraulic_radius=Rh[i],
                              slope=Se[i])
            mpm_dict["Q (m3/s)"].append(Q[i])
            mpm_dict["Phi (-)"].append(section_mpm.phi)
            b = hec_df["Flow Area"][i] / h[i]
            mpm_dict["Qb (kg/s)"].append(section_mpm.add_dimensions(b))
    return pd.DataFrame(mpm_dict)


def plot_bedload_transport(mpm_results):
    # Extract unique river stations
    river_stations = mpm_results["River Sta"].unique()

    if len(river_stations) < 3:
        print("Not enough river stations available for plotting.")
        return

    # Choose the first 3 unique river stations
    selected_stations = river_stations[:3]

    # Plot for each river station
    for station in selected_stations:
        # Filter rows for the selected river station
        subset = mpm_results[mpm_results["River Sta"] == station]

        # Create a new figure for this river station
        plt.figure(figsize=(8, 5))
        plt.plot(subset["Q (m3/s)"], subset["Qb (kg/s)"], marker='o', label=f"River Sta: {station}")

        # Customize the plot
        plt.xlabel("Discharge Q (m³/s)")
        plt.ylabel("Bedload Transport Qb (kg/s)")
        plt.title(f"Bedload Transport vs Discharge\nRiver Station: {station}")
        plt.grid(True)
        plt.legend()

        # Show the plot for the current river station
        plt.tight_layout()
        plt.show()


@log_actions
def main():
    # Get characteristic grain size = D84
    D_char = get_char_grain_size(file_name=os.path.abspath("../..") + "\\grains.csv",
                                 D_char="D84")
    hec_file = os.path.abspath("..") + "{0}HEC-RAS{0}output.xlsx".format(os.sep)
    hec = HecSet(hec_file)
    logging.info(hec.hec_data.head())

    mpm_results = calculate_mpm(hec.hec_data, D_char)
    mpm_results.to_excel(os.path.abspath("..") + os.sep + "bed_load_mpm.xlsx")

    # Add background color to headers
    add_header_background_color((os.path.abspath("..") + os.sep + "bed_load_mpm.xlsx"))

    # Select and plot 3 profiles
    plot_bedload_transport(mpm_results)

if __name__ == '__main__':
    main()
